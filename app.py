from flask import Flask, request, jsonify, render_template
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/register', methods=['POST'])
def register():
    try:
        # Get form data
        name = request.form.get('name')
        email = request.form.get('email')

        # Define the path to your Excel file
        excel_path = 'registration_data.xlsx'

        # Load existing workbook or create a new one if it doesn't exist
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "Email"])  # Add header row

        ws = wb.active

        # Append data to the worksheet
        ws.append([name, email])

        # Save the workbook
        wb.save(excel_path)

        # Return the success page
        return render_template('success.html')

    except Exception as e:
        # Log the exception (optional)
        app.logger.error(f"Error during registration: {e}")
        return jsonify({'error': 'Internal Server Error'}), 500

if __name__ == '__main__':
    app.run(debug=True)
